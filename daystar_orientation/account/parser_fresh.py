# will be used to parse the data of freshmen
from openpyxl import load_workbook

def extract_freshman_info(file):
    """
    Used to extract freshman information
    """
    # loading the workbook
    wb_obj = load_workbook(filename=file)
    work_sheets = wb_obj.sheetnames
    work_sheet = wb_obj[work_sheets[0]]
    headers = [] # will hold the titles
    data = [] # will hold the data as a list of dictionaries

    # getting headers
    for id, row in enumerate(work_sheet.iter_rows(values_only=True)):
        # getting columns names
        if id == 0:
            headers = row
            continue
        data.append(
            {headers[i]:data for i, data in enumerate(row)}
        )  
    return data